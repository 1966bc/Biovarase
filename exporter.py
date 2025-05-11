# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------
# project:  biovarase
# authors:  1966bc
# mailto:   [giuseppecostanzi@gmail.com]
# modify:   ver MMXXV
# This script has been significantly enhanced through collaboration with a programming assistant aka Gemini.
# -----------------------------------------------------------------------------
""" This is the exporter module of Biovarase."""
import sys
import inspect
import datetime
import tempfile

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class Exporter:

    def __str__(self):
        return "class: {0}\nMRO: {1}".format(self.__class__.__name__,
                                             [x.__name__ for x in Exporter.__mro__])

    def create_workbook(self, title='Biovarase'):
        try:
            wb = openpyxl.Workbook()
            worksheet = wb.active
            worksheet.title = title
            return wb, worksheet
        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
            return None, None

    def save_and_launch(self, workbook, suffix='.xlsx'):
        try:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            tmp.close()
            path = tmp.name
            workbook.save(path)
            self.launch(path)
            return path
        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
            return None

    def get_controls(self):

        try:
            today = datetime.datetime.now()
            sql = "SELECT samples.sample,\
                           tests.description,\
                           dict_tests.code,\
                           batches.lot_number,\
                           strftime('%d-%m-%Y', expiration),\
                           controls.description,\
                           controls.reference,\
                           suppliers.description,\
                           batches.expiration,\
                           batches.description,\
                           batches.target\
                      FROM tests\
                      INNER JOIN dict_tests ON tests.test_id = dict_tests.test_id\
                      INNER JOIN samples ON dict_tests.sample_id  = samples.sample_id\
                      INNER JOIN batches ON dict_tests.dict_test_id = batches.dict_test_id\
                      INNER JOIN controls ON batches.control_id = controls.control_id\
                      INNER JOIN suppliers ON controls.supplier_id = suppliers.supplier_id\
                      WHERE tests.status =1\
                        AND dict_tests.status =1\
                        AND batches.status =1\
                      ORDER BY tests.description ASC;"

            rs = self.read(True, sql, ())

            workbook, worksheet = self.create_workbook('Biovarase')
            if not workbook or not worksheet:
                return  # Exit if workbook creation failed

            #set column widths
            worksheet.column_dimensions[chr(ord('A') + 0)].width = 10
            worksheet.column_dimensions[chr(ord('A') + 1)].width = 30
            worksheet.column_dimensions[chr(ord('A') + 5)].width = 25
            worksheet.column_dimensions[chr(ord('A') + 7)].width = 30
            worksheet.column_dimensions[chr(ord('A') + 8)].width = 30
            row_num = 1  #rows in openpyxl start from 1

            #header
            cols = ("Sample", "Test", "Code", "Batch",
                    "Expiration", "Description", "Target",
                    "Control", "Reference", "Supplier")

            font_bold = Font(bold=True, name='Arial')
            for col_num, text in enumerate(cols):
                cell = worksheet.cell(row=row_num, column=col_num + 1, value=text)
                cell.font = font_bold

            row_num += 1

            if rs:
                for i in rs:
                    try:
                        expiration_str = i[4]
                        formatted_expiration = datetime.datetime.strptime(i[8], "%Y-%m-%d").date()
                        diff_date = formatted_expiration - today.date()
                        days_difference = diff_date.days

                        #sample
                        worksheet.cell(row=row_num, column=1, value=i[0])
                        #test
                        worksheet.cell(row=row_num, column=2, value=i[1])
                        #test method code
                        worksheet.cell(row=row_num, column=3, value=i[2])
                        #lot_number
                        worksheet.cell(row=row_num, column=4, value=i[3])
                        #compute expiration date
                        expiration_cell = worksheet.cell(row=row_num, column=5, value=expiration_str)
                        if days_difference <= 0:
                            expiration_cell.fill = PatternFill(start_color='FFFF0000',
                                                            end_color='FFFF0000',
                                                            fill_type='solid')  #red
                        elif days_difference <= 15:
                            expiration_cell.fill = PatternFill(start_color='FFFFFF00',
                                                            end_color='FFFFFF00',
                                                            fill_type='solid')  #yellow
                        #batches.description
                        worksheet.cell(row=row_num, column=6, value=i[9])
                        #batches.target
                        worksheet.cell(row=row_num, column=7, value=i[10])
                        #controls.description
                        worksheet.cell(row=row_num, column=8, value=i[5])
                        #controls.reference
                        worksheet.cell(row=row_num, column=9, value=i[6])
                        #suppliers.description
                        worksheet.cell(row=row_num, column=10, value=i[7])

                        row_num += 1
                    except (ValueError, TypeError) as ve:
                        self.on_log(inspect.stack()[0][3], ve, type(ve), sys.modules[__name__])
                        continue  # Log and continue to the next iteration
                    except Exception as e:
                        self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                        break  # Log and exit the loop

            self.save_and_launch(workbook)

        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

    def get_counts(self, selected_date):

        try:
            sql = "SELECT dict_tests.dict_test_id,\
                          dict_tests.code,\
                          tests.description,\
                          samples.sample,\
                          categories.description,\
                          COUNT(results.batch_id)\
                   FROM tests\
                   INNER JOIN dict_tests ON tests.test_id = dict_tests.test_id\
                   INNER JOIN batches ON dict_tests.dict_test_id = batches.dict_test_id\
                   INNER JOIN categories ON dict_tests.category_id = categories.category_id\
                   INNER JOIN samples ON dict_tests.sample_id = samples.sample_id\
                   INNER JOIN sections ON dict_tests.section_id = sections.section_id\
                   INNER JOIN labs ON sections.lab_id = labs.lab_id\
                   INNER JOIN results ON batches.batch_id = results.batch_id\
                   WHERE tests.status=1\
                   AND dict_tests.status=1\
                   AND sections.section_id =?\
                   AND DATE(results.received) >=?\
                   AND results.is_delete=0\
                   GROUP BY dict_tests.dict_test_id\
                   ORDER BY tests.description;"

            args = (self.get_section_id(), selected_date[0])

            rs = self.read(True, sql, args)

            workbook, worksheet = self.create_workbook('Biovarase')
            if not workbook or not worksheet:
                return

            row_num = 1

            #header
            cols = ('Code', 'Test', 'Sample', 'Categories', 'Count')
            for col_num, text in enumerate(cols):
                worksheet.cell(row=row_num, column=col_num + 1, value=text)

            row_num += 1

            if rs:
                for i in rs:
                    worksheet.cell(row=row_num, column=1, value=i[1])
                    worksheet.cell(row=row_num, column=2, value=i[2])
                    worksheet.cell(row=row_num, column=3, value=i[3])
                    worksheet.cell(row=row_num, column=4, value=i[4])
                    worksheet.cell(row=row_num, column=5, value=i[5])
                    row_num += 1

            self.save_and_launch(workbook)

        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

    def get_notes(self, args):

        try:
            workbook, worksheet = self.create_workbook('Biovarase')
            if not workbook or not worksheet:
                return

            sql = "SELECT tests.description,\
                           batches.lot_number,\
                           batches.target,\
                           batches.sd,\
                           results.result,\
                           results.received,\
                           actions.action,\
                           notes.description,\
                           notes.modified,\
                           equipments.description,\
                           workstations.description,\
                           workstations.serial,\
                           labs.lab,\
                           sections.section\
                      FROM tests\
                      INNER JOIN dict_tests ON tests.test_id = dict_tests.test_id\
                      INNER JOIN batches ON dict_tests.dict_test_id = batches.dict_test_id\
                      INNER JOIN results ON batches.batch_id = results.batch_id\
                      INNER JOIN workstations ON results.workstation_id = workstations.workstation_id\
                      INNER JOIN equipments ON workstations.equipment_id = equipments.equipment_id\
                      INNER JOIN sections ON workstations.section_id = sections.section_id\
                      INNER JOIN labs ON sections.lab_id = labs.lab_id\
                      INNER JOIN notes ON results.result_id = notes.result_id\
                      INNER JOIN actions ON notes.action_id = actions.action_id\
                      WHERE DATE(results.received) >=?\
                        AND sections.section_id =?\
                        AND tests.status = 1\
                        AND batches.status = 1\
                        AND results.is_delete=0\
                      ORDER BY notes.modified DESC;"

            row_num = 1

            #header
            cols = ("Test", "Batch", "Target", "SD", "Result",
                    "Recived", "Action", "Description", "Modify",
                    "Instrument", "Workstation", "Serial", "Ward", "Section")

            font_bold = Font(bold=True, name='Arial')
            for col_num, text in enumerate(cols):
                cell = worksheet.cell(row=row_num, column=col_num + 1, value=text)
                cell.font = font_bold

            row_num += 1

            rs = self.read(True, sql, args)

            if rs:
                for i in rs:
                    try:
                        worksheet.cell(row=row_num, column=1, value=i[0])
                        worksheet.cell(row=row_num, column=2, value=i[1])
                        worksheet.cell(row=row_num, column=3, value=i[2])

                        value_to_write = None
                        if i[3] is not None:
                            value_to_write = round(i[3], 3)
                        worksheet.cell(row=row_num, column=4, value=value_to_write)

                        value_to_write = None
                        if i[4] is not None:
                            value_to_write = round(i[4], 2)
                        worksheet.cell(row=row_num, column=5, value=value_to_write)

                        value_to_write = None
                        if i[5] is not None:
                            value_to_write = i[5].strftime("%d-%m-%Y")

                        worksheet.cell(row=row_num, column=6, value=value_to_write)
                        worksheet.cell(row=row_num, column=7, value=i[6])
                        worksheet.cell(row=row_num, column=8, value=i[7])
                        worksheet.cell(row=row_num, column=9, value=i[8])
                        worksheet.cell(row=row_num, column=10, value=i[9])
                        worksheet.cell(row=row_num, column=11, value=i[10])
                        worksheet.cell(row=row_num, column=12, value=i[11])
                        worksheet.cell(row=row_num, column=13, value=i[12])
                        worksheet.cell(row=row_num, column=14, value=i[13])
                        row_num += 1
                    except (TypeError, ValueError) as ve:
                        self.on_log(inspect.stack()[0][3], ve, type(ve), sys.modules[__name__])
                        continue
                    except Exception as e:
                        self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                        break

            self.save_and_launch(workbook)

        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])




    def quick_data_analysis(self, selected_date):

        try:
            checked_tests = []
            mandatory_tests = self.get_mandatory()
            workbook, worksheet = self.create_workbook('Biovarase')
            if not workbook or not worksheet:
                return

            # initialize column widths from F to M
            column_widths = {'F': 0, 'G': 0, 'H': 0, 'I': 0, 'J': 0, 'K': 0, 'L': 0, 'M': 0}

            row_num = 1  # openpyxl uses 1-based indexing
            cols = ('Type', 'Test', 'Batch', 'Expiration', 'Instrument', 'Target', 'Result', 'avg',
                    'bias', 'SD', 'sd', 'cv', 'Wstg', 'Date', 'Category', 'Workstation', 'Control', 'Supplier',
                    'Mandatory')

            # head style
            header_font = Font(bold=True, color='000000')  # Black
            for col_num, title in enumerate(cols, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=title)
                cell.font = header_font
                col_letter = get_column_letter(col_num)
                if col_letter in ['A', 'B', 'E', 'R']:
                    if col_letter == 'A':
                        worksheet.column_dimensions[col_letter].width = 10
                    elif col_letter == 'B':
                        worksheet.column_dimensions[col_letter].width = 30
                    elif col_letter == 'E':
                        worksheet.column_dimensions[col_letter].width = 20
                    elif col_letter == 'R':
                        worksheet.column_dimensions[col_letter].width = 50

            row_num += 1

            sql_tests = "SELECT dict_tests.dict_test_id,\
                                     samples.sample,\
                                     tests.description,\
                                     categories.description\
                                FROM tests\
                                INNER JOIN dict_tests ON tests.test_id = dict_tests.test_id\
                                INNER JOIN categories ON dict_tests.category_id = categories.category_id\
                                INNER JOIN samples ON dict_tests.sample_id = samples.sample_id\
                                INNER JOIN sections ON dict_tests.section_id = sections.section_id\
                                INNER JOIN labs ON sections.lab_id = labs.lab_id\
                                INNER JOIN sites ON labs.site_id = sites.site_id\
                                WHERE labs.lab_id =?\
                                  AND tests.status=1\
                                  AND dict_tests.status=1\
                                ORDER BY tests.description;"

            try:
                related_ids = self.get_related_ids_by_section(self.get_section_id())
                args = (related_ids[3],)
                rs_tests_methods = self.read(True, sql_tests, args)
            except Exception as e:
                self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                return

            for test_method in rs_tests_methods:
                sql_batches = "SELECT batches.*,\
                                       strftime('%d-%m-%Y', expiration),\
                                       workstations.description,\
                                       equipments.description\
                                  FROM batches\
                                  INNER JOIN workstations ON batches.workstation_id = workstations.workstation_id\
                                  INNER JOIN equipments ON workstations.equipment_id = equipments.equipment_id\
                                  INNER JOIN sections ON workstations.section_id = sections.section_id\
                                  INNER JOIN labs ON labs.lab_id = sections.lab_id\
                                  WHERE batches.status =1\
                                    AND batches.dict_test_id =?\
                                    AND labs.lab_id =?;"

                try:
                    rs_batches = self.read(True, sql_batches, (test_method[0], related_ids[3]))
                except Exception as e:
                    self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                    continue

                for batch in rs_batches:
                    expiration_date = batch[5]
                    workstation = batch[16]
                    instrument = batch[17]

                    sql_results = "SELECT results.result_id,\
                                           ROUND(results.result,2),\
                                           received,\
                                           workstations.description,\
                                           workstations.workstation_id,\
                                           results.received\
                                      FROM results\
                                      INNER JOIN workstations ON results.workstation_id = workstations.workstation_id\
                                      WHERE results.batch_id =?\
                                        AND DATE(results.received) =?\
                                        AND workstations.workstation_id =?\
                                        AND results.status = 1\
                                        AND results.is_delete=0\
                                      ORDER BY results.received DESC;"

                    try:
                        args_results = (batch[0], selected_date[0], batch[3],)
                        rs_results = self.read(True, sql_results, args_results)
                    except Exception as e:
                        self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                        continue

                    sql_controls = "SELECT controls.description, suppliers.description\
                                      FROM controls\
                                      INNER JOIN suppliers ON controls.supplier_id = suppliers.supplier_id\
                                      WHERE control_id =?"

                    try:
                        rs_controls = self.read(False, sql_controls, (batch[1],))
                    except Exception as e:
                        self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                        rs_controls = None

                    if rs_results is not None:
                        for result in rs_results:
                            try:
                                series = self.get_series(batch[0], result[4], int(self.get_observations()), result[0])
                            except Exception as e:
                                self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                                continue

                            if len(series) > 9:
                                try:
                                    rule = self.get_westgard_violation_rule(batch[6], batch[7], series, batch,
                                                                            test_method)
                                except Exception as e:
                                    self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                                    rule = "Error"
                            else:
                                rule = "NED"

                            result_color = None

                            try:
                                if result:
                                    compute_cv = self.get_cv(series)
                                    compute_sd = self.get_sd(series)
                                    compute_avg = self.get_mean(series)
                                    target = float(batch[6])
                                    sd = float(batch[7])
                                    bias = self.get_bias(compute_avg, target)
                                    res = float(result[1])
                                    date = result[2].strftime("%d-%m-%Y %H:%M:%S")

                                    if res > target:
                                        if res > (target + (sd * 3)):
                                            result_color = "red"
                                        elif res > (target + (sd * 2)) and res < (target + (sd * 3)):
                                            result_color = "yellow"
                                    elif res < target:
                                        if res < (target - (sd * 3)):
                                            result_color = "red"
                                        elif res < (target - (sd * 2)) and res > (target - (sd * 3)):
                                            result_color = "yellow"

                                    try:
                                        formatted_expiration = datetime.datetime.strptime(batch[5], "%Y-%m-%d").date()
                                        received_date = result[2].date()
                                        diff_date = formatted_expiration - received_date
                                        x = diff_date.days
                                    except (ValueError, TypeError) as ve:
                                        self.on_log(inspect.stack()[0][3], ve, type(ve), sys.modules[__name__])
                                        expiration_display = expiration_date
                                        expiration_fill = None
                                    else:
                                        expiration_display = formatted_expiration.strftime("%d-%m-%Y")
                                        expiration_fill = None
                                        if x <= 0:
                                            expiration_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000",
                                                                            fill_type="solid")  # red
                                            expiration_display = f"{expiration_display} (Expired)"
                                        elif x <= 15:
                                            expiration_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00",
                                                                            fill_type="solid")  # yellow
                                            expiration_display = f"{expiration_display} (Expires in {x} days)"

                                    row_values = [
                                        test_method[1],  # Type
                                        test_method[2],  # Test
                                        batch[4],  # Batch
                                        expiration_display,  # Expiration
                                        instrument,  # Instrument
                                        target,  # Target
                                        res,  # Result
                                        compute_avg,  # avg
                                        bias,  # bias
                                        sd,  # SD
                                        compute_sd,  # sd
                                        compute_cv,  # cv
                                        rule,  # Wstg (Westgard Rule)
                                        date,  # Date
                                        test_method[3],  # Category
                                        batch[18],  # Workstation
                                        rs_controls[0] if rs_controls and len(rs_controls) > 0 else "",  # Control
                                        rs_controls[1] if rs_controls and len(rs_controls) > 1 else "",  # Supplier
                                        "Yes" if test_method[1] in mandatory_tests else "No" # Mandatory
                                    ]
                                    for col_num, value in enumerate(row_values, 1):
                                        cell = worksheet.cell(row=row_num, column=col_num, value=value)
                                        col_letter = get_column_letter(col_num)
                                        if col_letter in column_widths and value is not None:
                                            current_width = len(str(value))
                                            if current_width > column_widths[col_letter]:
                                                column_widths[col_letter] = current_width
                                        if col_num == 4 and expiration_fill:  # expiration col
                                            cell.fill = expiration_fill
                                        if col_num == 7 and result_color:  # result col
                                            fill = PatternFill(start_color=self._convert_color(result_color),
                                                                end_color=self._convert_color(result_color),
                                                                fill_type="solid")
                                            cell.fill = fill
                                        if col_num == 13 and rule not in ('Accept', 'NED'):  # wstg col
                                            fill = PatternFill(start_color="FF0000FF", end_color="FF0000FF",
                                                                fill_type="solid")  # Blue
                                            cell.fill = fill

                                    checked_tests.append(test_method[1])
                                    row_num += 1

                            except Exception as e:
                                self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width + 2  # add padding

        for test in mandatory_tests:
            if test not in checked_tests:
                cell = worksheet.cell(row=row_num, column=19, value=test)  # mandatory field index 19
                red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                cell.fill = red_fill
                row_num += 1

        self.save_and_launch(workbook)

    def get_analitical_goals(self, limit, rs):
    
        try:
            workbook, worksheet = self.create_workbook('Biovarase')
            if not workbook or not worksheet:
                return

            column_widths = [6, 20, 8, 20, 6, 6, 6, 6, 6, 6, 6, 6, 6, 6, 6, 6, 6, 10, 25]
            max_width_a = 6
            max_width_c = 8
            max_width_e_q = 6
            
            # header
            cols = ('T', 'Analyte', 'Batch', 'Expiration Date', 'Target', 'AVG',
                    'CVa', 'CVw', 'CVb', 'Imp%', 'Bias%', 'TEa%', 'CVt', 'k imp',
                    'k bias', 'TE%', 'Drc%', 'Records', 'Workstation')
            font_bold = Font(bold=True, name='Arial')
            worksheet.append(list(cols))

            # set bold 
            for cell in worksheet[1]:
                try:
                    cell.font = font_bold
                except Exception as e:
                    self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

            for col_num, text in enumerate(cols):
                column_widths[col_num] = max(column_widths[col_num], len(str(text)) + 2)
                if col_num == 0:
                    column_widths[col_num] = min(column_widths[col_num], max_width_a + 2)
                elif col_num == 2:
                    column_widths[col_num] = min(column_widths[col_num], max_width_c + 2)
                elif 4 <= col_num <= 16:
                    column_widths[col_num] = min(column_widths[col_num], max_width_e_q + 2)

            row_num = 2

            for i in rs:
                try:
                    workstation_id = i[12] if len(i) > 12 else None
                    series = self.get_series(i[0], workstation_id, limit)

                    if len(series) > 5:
                        try:
                            cva = self.get_cv(series)
                            sd = self.get_sd(series)
                            avg = self.get_mean(series)
                            cvw = i[6]
                            cvb = i[7]
                            target = float(i[5])

                            formula_imp = self.get_formula_imp(row_num)
                            formula_bias = self.get_formula_bias(row_num)
                            formula_eta = self.get_formula_eta(row_num)
                            formula_cvt = self.get_formula_cvt(row_num)
                            formula_k_imp_res = self.get_formula_k_imp(cva, cvw, row_num)
                            formula_k_bias_res = self.get_formula_k_bias(avg, target, cvw, cva, row_num)
                            tea_tes_comparison_res = self.get_tea_tes_comparision(avg, target, cvw, cvb, sd, cva)
                            formula_drc = self.get_formula_drc(row_num)
                        except Exception as e:
                            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                            continue  # Skip to the next row in 'rs' if a calculation fails

                        workstation_description = None
                        if len(i) > 12:
                            try:
                                rs_workstation = self.read(False, "SELECT description FROM workstations WHERE workstation_id =?", (i[12],))
                                workstation_description = rs_workstation[0] if rs_workstation else None
                            except Exception as e:
                                self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                                workstation_description = None

                        row_data = [
                            i[1],
                            i[2],
                            str(i[3]),
                            str(i[4]),
                            target,
                            avg,
                            float(cva) if cva is not None else None,
                            float(cvw) if cvw is not None else None,
                            float(cvb) if cvb is not None else None,
                            f'=ROUND({formula_imp},2)',
                            f'=ROUND({formula_bias},2)',
                            f'=ROUND({formula_eta},2)',
                            f'=ROUND({formula_cvt},2)',
                            f'=ROUND({formula_k_imp_res[0] if formula_k_imp_res else "0"},2)',
                            f'=ROUND({formula_k_bias_res[0] if formula_k_bias_res else "0"},2)',
                            tea_tes_comparison_res[0],
                            f'=ROUND({formula_drc},2)',
                            len(series),
                            workstation_description
                        ]
                        try:
                            worksheet.append(row_data)

                            # adjust column width
                            for col_num, value in enumerate(row_data):
                                width = len(str(value)) + 2
                                if col_num == 0:  # col A
                                    column_widths[col_num] = min(max(column_widths[col_num], width), max_width_a + 2)
                                elif col_num == 2:  # col C
                                    column_widths[col_num] = min(max(column_widths[col_num], width), max_width_c + 2)
                                elif 4 <= col_num <= 16:  # cols E,Q
                                    column_widths[col_num] = min(max(column_widths[col_num], width), max_width_e_q + 2)
                                else:
                                    column_widths[col_num] = max(column_widths[col_num], width)

                            # blue color for cva
                            if cva is not None and cva > self.get_imp(cvw):
                                cell = worksheet.cell(row=row_num, column=7)
                                cell.fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')

                            if formula_k_imp_res:
                                cell = worksheet.cell(row=row_num, column=14)
                                if formula_k_imp_res[1] == 'yellow':
                                    cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                                elif formula_k_imp_res[1] == 'red':
                                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                elif formula_k_imp_res[1] == 'green':
                                    cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

                            if formula_k_bias_res:
                                cell = worksheet.cell(row=row_num, column=15)
                                if formula_k_bias_res[1] == 'yellow':
                                    cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                                elif formula_k_bias_res[1] == 'red':
                                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                elif formula_k_bias_res[1] == 'green':
                                    cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

                            if tea_tes_comparison_res[1]:
                                try:
                                    cell = worksheet.cell(row=row_num, column=16)
                                    fill_color = self._convert_color(tea_tes_comparison_res[1])
                                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                                except Exception as e:
                                    self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

                            row_num += 1
                        except Exception as e:
                            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

                except Exception as e:
                    self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

        # set cols width
        try:
            for i, width in enumerate(column_widths):
                worksheet.column_dimensions[get_column_letter(i + 1)].width = width
        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

        try:
            self.save_and_launch(workbook)
        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])

    def get_formula_drc(self, row):
        """compute critical difference"""
        return f"ROUND((SQRT(POWER(G{row},2)+POWER(H{row},2))*2.77)*F{row}/100,2)"

    def get_formula_imp(self, row):
        return f"ROUND((H{row} * 0.5),2)"

    def get_formula_bias(self, row):
        return f"ROUND(SQRT(POWER(H{row},2)+POWER(I{row},2))*0.25,2)"

    def get_formula_eta(self, row):
        z_score = self.get_zscore()
        return f"ROUND(({z_score}*J{row})+K{row},2)"

    def get_formula_cvt(self, row):
        return f"ROUND(SQRT(POWER(G{row},2)+POWER(H{row},2)),2)"

    def get_formula_k_imp(self, cva, cvw, row):
        try:
            if cvw is not None and cvw != 0:
                try:
                    k = round((cva / cvw), 2)
                except ZeroDivisionError as zde:
                    self.on_log(inspect.stack()[0][3], zde, type(zde), sys.modules[__name__])
                    return None, None
                
                if k is not None:  # Proceed only if k was successfully calculated
                    if 0.25 <= k <= 0.50:
                        c = "yellow"
                    elif 0.50 <= k <= 0.75:
                        c = "FFFF00"
                    elif k > 0.75:
                        c = "red"
                    else:
                        c = "yellow"
                    f = f"ROUND(G{row}/H{row},2)"
                    return f, c
                else:
                    return None, None
            else:
                return "0", None
        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
            return None, None

    def get_formula_k_bias(self, avg, target, cvw, cva, row):
        """Return bias k 0.125,0.25,0.375"""
        try:
            try:
                cvt = self.get_cvt(cva, cvw)
            except Exception as e:
                self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
                return None, None
            
            if cvt is not None and cvt != 0:
                try:
                    k = round(self.get_bias(avg, target) / cvt, 2)
                except ZeroDivisionError as zde:
                    self.on_log(inspect.stack()[0][3], zde, type(zde), sys.modules[__name__])
                    return None, None
                
                if k is not None:
                    if 0.125 <= k <= 0.25:
                        c = "yellow"
                    elif 0.25 <= k <= 0.375:
                        c = "FFFF00"
                    elif k > 0.375:
                        c = "red"
                    else:
                        c = "yellow"
                    f = f"ROUND((((F{row}-E{row})/E{row})*100)/SQRT(POWER(H{row},2)+POWER(I{row},2)),2)"
                    return f, c
                else:
                    return None, None
            else:
                return "0", None
        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
            return None, None

    def get_cvt(self, cva, cvw):
        try:
            cvt = round(math.sqrt(cva**2 + cvw**2), 2)
            return cvt
        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
            return None

    def get_bias(self, avg, target):
        try:
            bias = abs(round(float((avg - target) / float(target)) * 100, 2))
            return bias
        except ZeroDivisionError as zde:
            self.on_log(inspect.stack()[0][3], zde, type(zde), sys.modules[__name__])
            return None
        except ValueError as ve:
            self.on_log(inspect.stack()[0][3], ve, type(ve), sys.modules[__name__])
            return None
        except Exception as e:
            self.on_log(inspect.stack()[0][3], e, type(e), sys.modules[__name__])
            return None

    def _convert_color(self, color_name):
        color_map = {'red': 'FFFF0000', 'yellow': 'FFFFFF00', 'blue': 'FF0000FF', 'green': 'FF00FF00'}
        return color_map.get(color_name.lower(), 'FFFFFFFF')

def main():

    foo = Exporter()
    print(foo)
    input('end')

if __name__ == "__main__":
    main()
